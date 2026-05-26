import streamlit as st
import pandas as pd
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

from utils.database import read_sheet
from utils.formatters import formatar_data

# ---------------------------------------------------------------------------
# Titulo
# ---------------------------------------------------------------------------
st.header("Exportar Dados")

# ---------------------------------------------------------------------------
# Funcao auxiliar: formatar planilha Excel
# ---------------------------------------------------------------------------

def formatar_worksheet(ws, df: pd.DataFrame, colunas_peso: list[str] | None = None):
    """Escreve um DataFrame no worksheet e aplica formatacao."""
    if colunas_peso is None:
        colunas_peso = []

    # Cabecalhos
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            col_name = headers[col_idx - 1]
            if col_name in colunas_peso:
                try:
                    cell.value = float(value) if value != "" else 0
                    cell.number_format = '#,##0.0'
                except (ValueError, TypeError):
                    pass

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(ws.max_row + 1, 502)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_length = max(max_length, len(str(cell_val)))
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def gerar_excel_unico(df: pd.DataFrame, sheet_name: str, colunas_peso: list[str] | None = None) -> bytes:
    """Gera um arquivo Excel com uma unica aba."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    formatar_worksheet(ws, df, colunas_peso)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_excel_multi(sheets: dict[str, tuple[pd.DataFrame, list[str]]]) -> bytes:
    """Gera um arquivo Excel com multiplas abas.
    sheets: {nome_aba: (dataframe, colunas_peso)}
    """
    wb = Workbook()
    # Remover aba padrao
    wb.remove(wb.active)

    for nome_aba, (df, colunas_peso) in sheets.items():
        ws = wb.create_sheet(title=nome_aba[:31])  # Excel limita 31 chars
        formatar_worksheet(ws, df, colunas_peso)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Configuracoes de exportacao
# ---------------------------------------------------------------------------
MODULOS = [
    "Producao Lavacao",
    "Producao Extrusao",
    "Qualidade",
    "Estoque",
    "Romaneios",
    "Todos",
]

modulo = st.selectbox("Modulo para exportar", options=MODULOS)

col_d1, col_d2 = st.columns(2)
with col_d1:
    data_inicio = st.date_input(
        "Data Inicio",
        value=date.today() - timedelta(days=30),
        key="exp_data_inicio",
    )
with col_d2:
    data_fim = st.date_input(
        "Data Fim", value=date.today(), key="exp_data_fim"
    )

st.divider()


# ---------------------------------------------------------------------------
# Funcoes de carregamento por modulo
# ---------------------------------------------------------------------------

def carregar_producao_lavacao() -> pd.DataFrame:
    try:
        df = read_sheet("producao_lavacao")
    except Exception:
        return pd.DataFrame()
    if df.empty:
        return df
    df["data_dt"] = pd.to_datetime(df["data"], errors="coerce").dt.date
    df = df[(df["data_dt"] >= data_inicio) & (df["data_dt"] <= data_fim)]
    df = df.drop(columns=["data_dt", "id", "created_at"], errors="ignore")
    return df.sort_values("data", ascending=False) if not df.empty else df


def carregar_producao_extrusao() -> pd.DataFrame:
    try:
        df = read_sheet("producao_extrusao")
    except Exception:
        return pd.DataFrame()
    if df.empty:
        return df
    df["data_dt"] = pd.to_datetime(df["data"], errors="coerce").dt.date
    df = df[(df["data_dt"] >= data_inicio) & (df["data_dt"] <= data_fim)]
    df = df.drop(columns=["data_dt", "id", "created_at"], errors="ignore")
    return df.sort_values("data", ascending=False) if not df.empty else df


def carregar_qualidade() -> pd.DataFrame:
    try:
        df = read_sheet("qualidade")
    except Exception:
        return pd.DataFrame()
    if df.empty:
        return df
    df["data_analise_dt"] = pd.to_datetime(
        df["data_analise"], errors="coerce"
    ).dt.date
    df = df[
        (df["data_analise_dt"] >= data_inicio) & (df["data_analise_dt"] <= data_fim)
    ]
    df = df.drop(columns=["data_analise_dt", "id", "created_at"], errors="ignore")
    return (
        df.sort_values("data_analise", ascending=False) if not df.empty else df
    )


def carregar_estoque() -> pd.DataFrame:
    try:
        df_ext = read_sheet("producao_extrusao")
    except Exception:
        return pd.DataFrame()
    if df_ext.empty:
        return df_ext

    df_disp = df_ext[df_ext["status"] == "disponivel"].copy()
    if df_disp.empty:
        return pd.DataFrame()

    df_disp["peso_kg"] = pd.to_numeric(df_disp["peso_kg"], errors="coerce").fillna(0)

    try:
        df_qual = read_sheet("qualidade")
    except Exception:
        df_qual = pd.DataFrame()

    if not df_qual.empty:
        df_qual_unico = df_qual.drop_duplicates(subset="codigo_lote", keep="last")
        df_estoque = df_disp.merge(
            df_qual_unico[["codigo_lote", "grade", "cor", "local_estoque"]],
            on="codigo_lote",
            how="left",
        )
    else:
        df_estoque = df_disp.copy()
        df_estoque["grade"] = ""
        df_estoque["cor"] = ""
        df_estoque["local_estoque"] = ""

    df_estoque["data_dt"] = pd.to_datetime(
        df_estoque["data"], errors="coerce"
    ).dt.date
    df_estoque["dias_em_estoque"] = df_estoque["data_dt"].apply(
        lambda d: (date.today() - d).days if pd.notna(d) else 0
    )

    df_estoque = df_estoque[
        (df_estoque["data_dt"] >= data_inicio) & (df_estoque["data_dt"] <= data_fim)
    ]

    colunas = [
        "codigo_lote", "data", "tipo_descricao", "extrusora", "peso_kg",
        "grade", "cor", "local_estoque", "dias_em_estoque",
    ]
    colunas_presentes = [c for c in colunas if c in df_estoque.columns]
    df_estoque = df_estoque[colunas_presentes]
    return (
        df_estoque.sort_values("data", ascending=False) if not df_estoque.empty else df_estoque
    )


def carregar_romaneios() -> pd.DataFrame:
    try:
        df = read_sheet("romaneio")
    except Exception:
        return pd.DataFrame()
    if df.empty:
        return df
    df["data_dt"] = pd.to_datetime(df["data"], errors="coerce").dt.date
    df = df[(df["data_dt"] >= data_inicio) & (df["data_dt"] <= data_fim)]
    df = df.drop(columns=["data_dt", "id", "created_at"], errors="ignore")
    return df.sort_values("data", ascending=False) if not df.empty else df


# ---------------------------------------------------------------------------
# Mapeamento de modulos
# ---------------------------------------------------------------------------
MAPA_MODULOS = {
    "Producao Lavacao": {
        "func": carregar_producao_lavacao,
        "sheet_name": "Producao Lavacao",
        "colunas_peso": ["peso_kg", "perda_lixo_kg", "perda_papelao_kg",
                         "perda_plastico_colorido_kg", "perda_total_kg"],
        "file_name": "producao_lavacao",
    },
    "Producao Extrusao": {
        "func": carregar_producao_extrusao,
        "sheet_name": "Producao Extrusao",
        "colunas_peso": ["peso_kg"],
        "file_name": "producao_extrusao",
    },
    "Qualidade": {
        "func": carregar_qualidade,
        "sheet_name": "Qualidade",
        "colunas_peso": ["mfi", "teor_cinzas", "densidade", "umidade"],
        "file_name": "qualidade",
    },
    "Estoque": {
        "func": carregar_estoque,
        "sheet_name": "Estoque",
        "colunas_peso": ["peso_kg"],
        "file_name": "estoque",
    },
    "Romaneios": {
        "func": carregar_romaneios,
        "sheet_name": "Romaneios",
        "colunas_peso": ["peso_total_kg"],
        "file_name": "romaneios",
    },
}

# ---------------------------------------------------------------------------
# Preview e Download
# ---------------------------------------------------------------------------
st.subheader("Pre-visualizacao")

if modulo == "Todos":
    # Carregar todos os modulos
    all_sheets = {}
    algum_dado = False
    for nome, config in MAPA_MODULOS.items():
        df = config["func"]()
        if not df.empty:
            algum_dado = True
            st.markdown(f"**{nome}** - {len(df)} registros")
            st.dataframe(df.head(10), use_container_width=True, hide_index=True)
            all_sheets[config["sheet_name"]] = (df, config["colunas_peso"])
        else:
            st.caption(f"{nome}: sem dados no periodo.")

    if algum_dado:
        excel_bytes = gerar_excel_multi(all_sheets)
        nome_arquivo = (
            f"resiban_todos_{data_inicio.strftime('%Y%m%d')}_"
            f"{data_fim.strftime('%Y%m%d')}.xlsx"
        )
        st.divider()
        st.download_button(
            label="Baixar Excel (Todos os Modulos)",
            data=excel_bytes,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    else:
        st.info("Nenhum dado encontrado no periodo selecionado.")

else:
    config = MAPA_MODULOS[modulo]
    df = config["func"]()

    if df.empty:
        st.info(f"Nenhum dado de '{modulo}' encontrado no periodo selecionado.")
    else:
        st.markdown(f"**{len(df)} registros encontrados**")
        st.dataframe(df, use_container_width=True, hide_index=True)

        excel_bytes = gerar_excel_unico(
            df, config["sheet_name"], config["colunas_peso"]
        )
        nome_arquivo = (
            f"resiban_{config['file_name']}_{data_inicio.strftime('%Y%m%d')}_"
            f"{data_fim.strftime('%Y%m%d')}.xlsx"
        )
        st.divider()
        st.download_button(
            label=f"Baixar Excel - {modulo}",
            data=excel_bytes,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
